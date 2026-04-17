from django.db.models import Avg, Count, Q, FloatField
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
import django_filters
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import FundManager, Fund, WorkflowTask, TaskComment, PlatformUser
from .serializers import (
    FundManagerListSerializer, FundManagerDetailSerializer,
    FundSerializer, WorkflowTaskSerializer, TaskCommentSerializer,
    OverviewStatsSerializer, VintageStatsSerializer,
)


# ──────────────────────────────────────────────
# Auth views  (no authentication required)
# ──────────────────────────────────────────────

@csrf_exempt
@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def login_view(request):
    email = request.data.get('email', '').strip().lower()
    password = request.data.get('password', '')

    if not email or not password:
        return Response({'error': 'Email and password are required.'}, status=400)

    try:
        user = PlatformUser.objects.get(email__iexact=email, is_active=True)
    except PlatformUser.DoesNotExist:
        return Response({'error': 'Invalid email or password.'}, status=401)

    if not user.check_password(password):
        return Response({'error': 'Invalid email or password.'}, status=401)

    user.last_login = timezone.now()
    user.save(update_fields=['last_login'])

    request.session['platform_user_id'] = user.id
    request.session['platform_user_email'] = user.email
    request.session.save()

    return Response({'id': user.id, 'email': user.email})


@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def logout_view(request):
    request.session.flush()
    return Response({'ok': True})


@api_view(['GET'])
@authentication_classes([])
@permission_classes([AllowAny])
def me_view(request):
    user_id = request.session.get('platform_user_id')
    if not user_id:
        return Response({'error': 'Not authenticated.'}, status=401)
    try:
        user = PlatformUser.objects.get(id=user_id, is_active=True)
        return Response({'id': user.id, 'email': user.email})
    except PlatformUser.DoesNotExist:
        request.session.flush()
        return Response({'error': 'Not authenticated.'}, status=401)


# ──────────────────────────────────────────────
# Filters
# ──────────────────────────────────────────────

class FundManagerFilter(django_filters.FilterSet):
    strategy = django_filters.CharFilter(lookup_expr='iexact')
    min_irr = django_filters.NumberFilter(field_name='avg_irr', lookup_expr='gte')
    max_irr = django_filters.NumberFilter(field_name='avg_irr', lookup_expr='lte')
    min_tvpi = django_filters.NumberFilter(field_name='avg_tvpi', lookup_expr='gte')
    min_dpi = django_filters.NumberFilter(field_name='avg_dpi', lookup_expr='gte')
    min_pb_score = django_filters.NumberFilter(field_name='pb_score', lookup_expr='gte')
    year_founded = django_filters.NumberFilter(field_name='year_founded')

    class Meta:
        model = FundManager
        fields = ['strategy', 'min_irr', 'max_irr', 'min_tvpi', 'min_dpi', 'min_pb_score', 'year_founded']


class FundFilter(django_filters.FilterSet):
    vintage = django_filters.NumberFilter()
    vintage_min = django_filters.NumberFilter(field_name='vintage', lookup_expr='gte')
    vintage_max = django_filters.NumberFilter(field_name='vintage', lookup_expr='lte')
    fund_type = django_filters.CharFilter(lookup_expr='iexact')
    strategy = django_filters.CharFilter(lookup_expr='iexact')
    quartile = django_filters.CharFilter(field_name='fund_quartile', lookup_expr='icontains')
    min_irr = django_filters.NumberFilter(field_name='irr', lookup_expr='gte')
    max_irr = django_filters.NumberFilter(field_name='irr', lookup_expr='lte')
    min_tvpi = django_filters.NumberFilter(field_name='tvpi', lookup_expr='gte')
    manager_name = django_filters.CharFilter(field_name='manager__name', lookup_expr='icontains')

    class Meta:
        model = Fund
        fields = ['vintage', 'fund_type', 'strategy', 'quartile', 'min_irr', 'max_irr', 'min_tvpi', 'manager_name']


# ──────────────────────────────────────────────
# ViewSets
# ──────────────────────────────────────────────

class FundManagerViewSet(viewsets.ModelViewSet):
    queryset = FundManager.objects.all()
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = FundManagerFilter
    search_fields = ['name', 'description']
    ordering_fields = ['avg_irr', 'avg_tvpi', 'avg_dpi', 'pb_score', 'aum_usd_m', 'fund_count', 'name']
    ordering = ['-avg_irr']

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return FundManagerDetailSerializer
        return FundManagerListSerializer

    @action(detail=False, methods=['get'])
    def top_performers(self, request):
        """Top N managers by a given metric"""
        metric = request.query_params.get('metric', 'avg_irr')
        n = int(request.query_params.get('n', 15))
        valid = ['avg_irr', 'avg_tvpi', 'avg_dpi', 'pb_score', 'aum_usd_m']
        if metric not in valid:
            return Response({'error': f'metric must be one of {valid}'}, status=400)
        qs = self.get_queryset().filter(**{f'{metric}__isnull': False}).order_by(f'-{metric}')[:n]
        return Response(FundManagerListSerializer(qs, many=True).data)

    @action(detail=False, methods=['get'])
    def overview_stats(self, request):
        """Aggregate stats for the overview dashboard, respecting filters"""
        qs = self.filter_queryset(self.get_queryset())
        fund_qs = Fund.objects.filter(manager__in=qs)

        irrs = list(qs.filter(avg_irr__isnull=False).values_list('avg_irr', flat=True))
        tvpis = list(qs.filter(avg_tvpi__isnull=False).values_list('avg_tvpi', flat=True))
        dpis = list(qs.filter(avg_dpi__isnull=False).values_list('avg_dpi', flat=True))
        aums = list(qs.filter(aum_usd_m__isnull=False).values_list('aum_usd_m', flat=True))

        def safe_avg(lst): return sum(lst)/len(lst) if lst else None
        def safe_median(lst):
            if not lst: return None
            s = sorted(lst)
            m = len(s)//2
            return s[m] if len(s)%2 else (s[m-1]+s[m])/2

        vintages = list(fund_qs.filter(vintage__isnull=False).values_list('vintage', flat=True).distinct().order_by('vintage'))

        data = {
            'total_managers': qs.count(),
            'managers_with_irr': qs.filter(avg_irr__isnull=False).count(),
            'total_funds': fund_qs.count(),
            'avg_irr': safe_avg(irrs),
            'median_irr': safe_median(irrs),
            'avg_tvpi': safe_avg(tvpis),
            'avg_dpi': safe_avg(dpis),
            'total_aum_usd_m': sum(aums) if aums else None,
            'vintage_range': [min(vintages), max(vintages)] if vintages else [],
        }
        return Response(data)

    @action(detail=False, methods=['get'])
    def vintage_stats(self, request):
        """IRR, TVPI, DPI aggregated by vintage year"""
        from django.db.models import Avg, Count
        stats = (
            Fund.objects
            .filter(vintage__isnull=False)
            .values('vintage')
            .annotate(
                avg_irr=Avg('irr'),
                avg_tvpi=Avg('tvpi'),
                avg_dpi=Avg('dpi'),
                fund_count=Count('id'),
            )
            .order_by('vintage')
        )
        return Response(list(stats))

    @action(detail=False, methods=['get'])
    def scatter_data(self, request):
        """Return manager-level data points for scatter plots"""
        qs = self.filter_queryset(self.get_queryset())
        data = list(qs.values(
            'name', 'strategy', 'avg_irr', 'avg_tvpi', 'avg_dpi', 'avg_rvpi',
            'pb_score', 'aum_usd_m', 'fund_count',
        ))
        return Response(data)


class FundViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Fund.objects.select_related('manager').all()
    serializer_class = FundSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = FundFilter
    search_fields = ['fund_name', 'manager__name', 'preferred_geography', 'preferred_industry']
    ordering_fields = ['irr', 'tvpi', 'dpi', 'vintage', 'fund_size', 'investments']
    ordering = ['-irr']


class WorkflowTaskViewSet(viewsets.ModelViewSet):
    queryset = WorkflowTask.objects.select_related('manager').prefetch_related('comments').all()
    serializer_class = WorkflowTaskSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'priority', 'manager']
    search_fields = ['title', 'description', 'assignee']
    ordering_fields = ['created_at', 'updated_at', 'due_date', 'priority']
    ordering = ['-created_at']

    @action(detail=True, methods=['post'])
    def add_comment(self, request, pk=None):
        task = self.get_object()
        serializer = TaskCommentSerializer(data={**request.data, 'task': task.id})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['patch'])
    def update_status(self, request, pk=None):
        task = self.get_object()
        new_status = request.data.get('status')
        valid = [c[0] for c in WorkflowTask.STATUS_CHOICES]
        if new_status not in valid:
            return Response({'error': f'status must be one of {valid}'}, status=400)
        task.status = new_status
        task.save(update_fields=['status', 'updated_at'])
        return Response(WorkflowTaskSerializer(task).data)


class TaskCommentViewSet(viewsets.ModelViewSet):
    queryset = TaskComment.objects.all()
    serializer_class = TaskCommentSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['task']


# ──────────────────────────────────────────────
# Excel Export
# ──────────────────────────────────────────────

@api_view(['GET'])
@authentication_classes([])
@permission_classes([AllowAny])
def export_funds_excel(request):
    """Export all fund details (with strategy) to an Excel file."""
    # Apply same filters as the FundViewSet
    qs = Fund.objects.select_related('manager').all()

    # Optional filters from query params
    strategy = request.query_params.get('strategy')
    fund_type = request.query_params.get('fund_type')
    vintage = request.query_params.get('vintage')
    quartile = request.query_params.get('quartile')
    min_irr = request.query_params.get('min_irr')

    if strategy:
        qs = qs.filter(strategy__iexact=strategy)
    if fund_type:
        qs = qs.filter(fund_type__iexact=fund_type)
    if vintage:
        qs = qs.filter(vintage=vintage)
    if quartile:
        qs = qs.filter(fund_quartile__icontains=quartile)
    if min_irr:
        qs = qs.filter(irr__gte=min_irr)

    qs = qs.order_by('-irr')

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fund Details'

    # Styles
    header_font = Font(bold=True, color='0A1628', size=11)
    header_fill = PatternFill(start_color='F0B429', end_color='F0B429', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

    # Headers
    headers = [
        'Fund ID', 'Manager', 'Fund Name', 'Vintage', 'Strategy',
        'Fund Type', 'Fund Size ($M)', '# Investments',
        'IRR (%)', 'TVPI (x)', 'DPI (x)', 'RVPI (x)',
        'Quartile', 'IRR Benchmark', 'TVPI Benchmark', 'DPI Benchmark',
        'As Of Quarter', 'As Of Year',
        'Preferred Geography', 'Preferred Industry',
    ]

    col_widths = [18, 30, 35, 10, 14, 22, 16, 14, 10, 10, 10, 10, 22, 14, 14, 14, 14, 12, 30, 30]

    ws.row_dimensions[1].height = 30
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, fund in enumerate(qs, start=2):
        row_fill = alt_fill if row_idx % 2 == 0 else None
        values = [
            fund.fund_id,
            fund.manager.name,
            fund.fund_name,
            fund.vintage,
            fund.strategy or '',
            fund.fund_type or '',
            round(fund.fund_size, 1) if fund.fund_size is not None else '',
            int(fund.investments) if fund.investments is not None else '',
            round(fund.irr, 1) if fund.irr is not None else '',
            round(fund.tvpi, 2) if fund.tvpi is not None else '',
            round(fund.dpi, 2) if fund.dpi is not None else '',
            round(fund.rvpi, 2) if fund.rvpi is not None else '',
            fund.fund_quartile or '',
            round(fund.irr_benchmark, 1) if fund.irr_benchmark is not None else '',
            round(fund.tvpi_benchmark, 2) if fund.tvpi_benchmark is not None else '',
            round(fund.dpi_benchmark, 2) if fund.dpi_benchmark is not None else '',
            fund.as_of_quarter or '',
            fund.as_of_year or '',
            fund.preferred_geography or '',
            fund.preferred_industry or '',
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if row_fill:
                cell.fill = row_fill

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto filter
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    # Stream response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="fund_details_export.xlsx"'
    wb.save(response)
    return response
